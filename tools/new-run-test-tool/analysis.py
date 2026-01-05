#!/usr/bin/env python3
"""
Enhanced Run Test Analyzer - Memory-Efficient Analysis Engine
Finds valid 72-hour reliability test windows from SCADA data per M01 contractual requirements.

Key Features:
- Per-WTG sequential processing to minimize memory usage
- Sliding window + prefix sums for O(n) window search
- Fault code 3-strike detection
- 50% rated power cooling verification
- 24h cumulative nominal power tracking
- Availability >= 96% validation
"""

import io
import json
import re
import gc
import sys
import numpy as np
import pandas as pd
from collections import Counter
from datetime import datetime, timedelta

# Global storage for results
output_excel_bytes = None
cached_analysis_data = None


def log(msg):
    """Print and flush for real-time output in browser."""
    print(msg)
    sys.stdout.flush()


# =============================================================================
# COLUMN DETECTION
# =============================================================================

TIMESTAMP_CANDIDATES = ['PCTimeStamp', 'Timestamp', 'TimeStamp', 'DateTime', 'Datetime', 'Time', 'Date']

# Column suffix patterns for WTG data (with numbered suffix like "(1)")
# Order matters - first match wins, so prefer instantaneous power over cumulative
COLUMN_PATTERNS = {
    'category': ['1_Report Category', 'Report Category'],
    'state': ['System States TurbineState'],
    'error_code': ['1_Status Error Code', 'Status Error Code'],
    'power': ['Grid Production Power Avg.', 'Power, Average', 'Power Avg'],  # NOT 'Total Active power' - that's cumulative energy
    'wind_speed': ['Ambient WindSpeed Avg.', 'Wind speed, Average'],
    'air_density': ['Ambient Airdensity AirDensityAvg Avg', 'Airdensity']
}

# Alarm column candidates
ALARM_TS_CANDIDATES = ['Detected', 'Event time', 'Occurred', 'Alarm time', 'Start time', 'Timestamp']
ALARM_UNIT_CANDIDATES = ['Unit', 'WTG', 'Turbine', 'Turbine ID', 'Unit ID']
ALARM_CODE_CANDIDATES = ['Code', 'Alarm Code', 'Event Code', 'Fault Code']
ALARM_DESC_CANDIDATES = ['Description', 'Alarm Description', 'Event Description', 'Message']
ALARM_TYPE_CANDIDATES = ['Event type', 'Type', 'EventType', 'Alarm Type']

# Alarm vs Warning classification (alarms take priority over warnings)
ALARM_EVENT_TYPES = {'alarm', 'fault', 'error', 'critical', 'failure'}
WARNING_EVENT_TYPES = {'warning', 'caution', 'notice', 'info'}

# Category classification
ACTIVE_CATEGORIES = {'normal operation'}
BLOCKED_CATEGORIES = {'manufacturer', 'unscheduled maintenance'}
ALLOWED_WINDOW_CATEGORIES = {'normal operation', 'scheduled maintenance', 'owner', 'environmental', 'utility'}


def _clean_colname(name: str) -> str:
    """Clean column name: strip whitespace, remove trailing (N), normalize spaces."""
    s = str(name).strip()
    s = re.sub(r"\s*\(\d+\)\s*$", "", s)  # Remove trailing (123)
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_category(cat: str) -> str:
    """Normalize category string for comparison."""
    if pd.isna(cat):
        return ''
    return str(cat).strip().lower()


def find_timestamp_column(columns: list) -> str:
    """Find the timestamp column from a list of column names."""
    clean_cols = {_clean_colname(c): c for c in columns}
    for candidate in TIMESTAMP_CANDIDATES:
        if candidate in clean_cols:
            return clean_cols[candidate]
        for clean, orig in clean_cols.items():
            if candidate.lower() == clean.lower():
                return orig
    return None


def find_wtg_column(columns: list, wtg: str, pattern_key: str) -> str:
    """Find column for a specific WTG and pattern type."""
    patterns = COLUMN_PATTERNS.get(pattern_key, [])
    wtg_upper = wtg.upper()
    
    for col in columns:
        clean = _clean_colname(col)
        # Check if column starts with WTG prefix
        if not (clean.upper().startswith(wtg_upper + '_') or 
                clean.upper().startswith(wtg_upper + ' ') or
                clean.upper().startswith(wtg_upper + '-')):
            continue
        # Check if any pattern matches
        for pattern in patterns:
            if pattern.lower() in clean.lower():
                return col
    return None


def detect_wtgs(columns: list) -> list:
    """Detect all WTG prefixes from column names."""
    wtgs = set()
    pattern = re.compile(r'^([A-Z]\d{2})[_\s-]', re.IGNORECASE)
    for col in columns:
        clean = _clean_colname(col)
        match = pattern.match(clean)
        if match:
            wtgs.add(match.group(1).upper())
    return sorted(list(wtgs))


def find_alarm_column(columns: list, candidates: list) -> str:
    """Find an alarm log column from candidates."""
    clean_map = {_clean_colname(c).lower(): c for c in columns}
    for candidate in candidates:
        cand_lower = candidate.lower()
        if cand_lower in clean_map:
            return clean_map[cand_lower]
        for clean, orig in clean_map.items():
            if cand_lower in clean:
                return orig
    return None


# =============================================================================
# POWER CURVE PARSING
# =============================================================================

def _is_number(val: str) -> bool:
    try:
        float(val)
        return True
    except:
        return False


def parse_power_curve_text(text: str) -> pd.DataFrame:
    """Parse a pasted power curve table (wind speed rows, air density columns)."""
    if not text or not text.strip():
        raise ValueError('Power curve text is empty.')
    raw_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not raw_lines:
        raise ValueError('Power curve text is empty after trimming.')

    def _tokens(line: str) -> list:
        return [t for t in re.split(r"[\s,\t]+", line.strip()) if t]

    def _is_air_density_like(val: float) -> bool:
        """Check if value looks like an air density (0.8-1.4 kg/m³)."""
        return 0.8 <= val <= 1.4

    # Find header row with air density values
    # Skip title rows like "Air density [kg/m3]"
    header_idx = None
    header_tokens = []
    
    for idx, ln in enumerate(raw_lines[:10]):  # Check first 10 lines
        toks = _tokens(ln)
        if not toks:
            continue
        
        # Count numeric tokens that look like air densities
        numeric_tokens = []
        for t in toks:
            if _is_number(t):
                numeric_tokens.append(float(t))
        
        ad_like_count = sum(1 for v in numeric_tokens if _is_air_density_like(v))
        
        # Header should have multiple air-density-like values (>=3) or many numeric values (>=5)
        if ad_like_count >= 3 or len(numeric_tokens) >= 5:
            header_idx = idx
            header_tokens = toks
            break
        
        # Also accept if line contains 'ws' or 'wind' with numbers
        joined = ' '.join(toks).lower()
        if ('ws' in joined or 'wind' in joined) and len(numeric_tokens) >= 2:
            header_idx = idx
            header_tokens = toks
            break
    
    if header_idx is None:
        # Fallback: try first row as header
        header_idx = 0
        header_tokens = _tokens(raw_lines[0])
    
    num_start = next((i for i, t in enumerate(header_tokens) if _is_number(t)), None)
    if num_start is None:
        raise ValueError('Header row missing numeric air density values.')
    
    first_col = ' '.join(header_tokens[:num_start]).strip() or 'WS'
    density_values = [float(t) for t in header_tokens[num_start:]]
    expected_cols = 1 + len(density_values)

    rows = []
    i = header_idx + 1
    while i < len(raw_lines):
        toks = _tokens(raw_lines[i])
        if not toks:
            i += 1
            continue
        if len(toks) == 1 and _is_number(toks[0]) and i + 1 < len(raw_lines):
            nxt = _tokens(raw_lines[i + 1])
            if len(nxt) == len(density_values) and all(_is_number(t) for t in nxt):
                row = [float(toks[0])] + [float(x) for x in nxt]
                rows.append(row)
                i += 2
                continue
        if _is_number(toks[0]) and len(toks) >= 2:
            try:
                row = [float(toks[0])] + [float(x) for x in toks[1:expected_cols]]
                if len(row) == expected_cols:
                    rows.append(row)
            except:
                pass
        i += 1

    if not rows:
        raise ValueError('No power curve rows parsed; check pasted format.')

    col_names = [first_col] + [str(d) for d in density_values]
    return pd.DataFrame(rows, columns=col_names)


def prepare_power_curve(curve_df: pd.DataFrame):
    """Prepare power curve arrays for interpolation."""
    ws_vals = pd.to_numeric(curve_df.iloc[:, 0], errors='coerce').to_numpy()
    ad_vals = np.array([float(c) for c in curve_df.columns[1:]], dtype=float)
    grid = curve_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').to_numpy(dtype=float)
    
    # Sort air density columns (required for interpolation)
    sort_idx = np.argsort(ad_vals)
    ad_vals = ad_vals[sort_idx]
    grid = grid[:, sort_idx]
    
    # Sort wind speed rows (required for interpolation)
    ws_sort_idx = np.argsort(ws_vals)
    ws_vals = ws_vals[ws_sort_idx]
    grid = grid[ws_sort_idx, :]
    
    if ws_vals.size == 0 or ad_vals.size == 0:
        raise ValueError('Power curve must include wind speed and air density columns.')
    return ws_vals, ad_vals, grid


def interpolate_power_curve(curve_arrays, wind_speed: float, air_density: float) -> float:
    """Bilinear interpolation of power curve."""
    ws_vals, ad_vals, grid = curve_arrays
    if not np.isfinite(wind_speed) or not np.isfinite(air_density):
        return float('nan')
    
    ws_clamped = float(np.clip(wind_speed, ws_vals.min(), ws_vals.max()))
    ad_clamped = float(np.clip(air_density, ad_vals.min(), ad_vals.max()))

    i = np.searchsorted(ws_vals, ws_clamped)
    j = np.searchsorted(ad_vals, ad_clamped)
    i0 = max(i - 1, 0)
    i1 = min(i, len(ws_vals) - 1)
    j0 = max(j - 1, 0)
    j1 = min(j, len(ad_vals) - 1)
    
    ws0, ws1 = ws_vals[i0], ws_vals[i1]
    ad0, ad1 = ad_vals[j0], ad_vals[j1]
    q11, q12, q21, q22 = grid[i0, j0], grid[i0, j1], grid[i1, j0], grid[i1, j1]

    if ws1 == ws0 and ad1 == ad0:
        return float(q11)
    if ws1 == ws0:
        return float(q11 + (q12 - q11) * (ad_clamped - ad0) / (ad1 - ad0 + 1e-9))
    if ad1 == ad0:
        return float(q11 + (q21 - q11) * (ws_clamped - ws0) / (ws1 - ws0 + 1e-9))
    
    denom = (ws1 - ws0) * (ad1 - ad0)
    w11 = (ws1 - ws_clamped) * (ad1 - ad_clamped)
    w12 = (ws1 - ws_clamped) * (ad_clamped - ad0)
    w21 = (ws_clamped - ws0) * (ad1 - ad_clamped)
    w22 = (ws_clamped - ws0) * (ad_clamped - ad0)
    return float((q11 * w11 + q12 * w12 + q21 * w21 + q22 * w22) / denom)


def compute_expected_power(wind_series: pd.Series, air_series: pd.Series, curve_arrays) -> np.ndarray:
    """Compute expected power for each row using power curve interpolation."""
    if curve_arrays is None:
        return np.full(len(wind_series), np.nan)
    
    result = np.empty(len(wind_series), dtype=np.float32)
    ws = wind_series.to_numpy()
    ad = air_series.to_numpy()
    
    for i in range(len(ws)):
        result[i] = interpolate_power_curve(curve_arrays, ws[i], ad[i])
    
    return result


# =============================================================================
# ALARM LOG PROCESSING
# =============================================================================

def parse_alarm_log(alarm_data: dict, columns: list) -> pd.DataFrame:
    """Parse alarm log data into DataFrame with standardized columns."""
    df = pd.DataFrame(alarm_data)
    
    # Find key columns
    ts_col = find_alarm_column(columns, ALARM_TS_CANDIDATES)
    unit_col = find_alarm_column(columns, ALARM_UNIT_CANDIDATES)
    code_col = find_alarm_column(columns, ALARM_CODE_CANDIDATES)
    desc_col = find_alarm_column(columns, ALARM_DESC_CANDIDATES)
    type_col = find_alarm_column(columns, ALARM_TYPE_CANDIDATES)
    
    if ts_col is None or unit_col is None:
        raise ValueError('Alarm log must have timestamp and unit columns')
    
    # Standardize column names
    result = pd.DataFrame()
    result['timestamp'] = pd.to_datetime(df[ts_col], errors='coerce')
    result['unit'] = df[unit_col].astype(str).str.upper().str.strip()
    result['code'] = df[code_col].astype(str) if code_col else ''
    result['description'] = df[desc_col].astype(str) if desc_col else ''
    result['event_type'] = df[type_col].astype(str) if type_col else ''
    
    return result.dropna(subset=['timestamp'])


def get_alarms_for_wtg(alarm_df: pd.DataFrame, wtg: str) -> pd.DataFrame:
    """Filter alarm log for a specific WTG."""
    return alarm_df[alarm_df['unit'] == wtg.upper()].copy()


def count_fault_codes_in_window(alarm_df: pd.DataFrame, start_ts, end_ts) -> dict:
    """Count occurrences of each fault code within a time window."""
    mask = (alarm_df['timestamp'] >= start_ts) & (alarm_df['timestamp'] <= end_ts)
    window_alarms = alarm_df[mask]
    return dict(Counter(window_alarms['code'].tolist()))


def count_alarms_and_warnings(alarm_df: pd.DataFrame, start_ts, end_ts) -> tuple:
    """
    Count alarms and warnings separately within a time window.
    Returns (alarm_count, warning_count).
    Alarms are events with type matching ALARM_EVENT_TYPES.
    Warnings are events with type matching WARNING_EVENT_TYPES.
    Unclassified events default to alarms for safety.
    """
    if alarm_df is None or len(alarm_df) == 0:
        return 0, 0
    
    mask = (alarm_df['timestamp'] >= start_ts) & (alarm_df['timestamp'] <= end_ts)
    window_events = alarm_df[mask]
    
    if len(window_events) == 0:
        return 0, 0
    
    alarm_count = 0
    warning_count = 0
    
    for _, row in window_events.iterrows():
        event_type = str(row.get('event_type', '')).strip().lower()
        # Check if it's a warning
        if any(w in event_type for w in WARNING_EVENT_TYPES):
            warning_count += 1
        else:
            # Default to alarm (including unclassified events)
            alarm_count += 1
    
    return alarm_count, warning_count


def check_three_strike_rule(fault_counts: dict) -> tuple:
    """Check if any fault code appears 3+ times. Returns (passes, violating_codes)."""
    violating = {code: count for code, count in fault_counts.items() if count >= 3}
    return len(violating) == 0, violating


# =============================================================================
# WINDOW SEARCH ALGORITHM
# =============================================================================

def extract_wtg_columns(scada_data: dict, columns: list, wtg: str, ts_col: str) -> dict:
    """
    Extract only the columns needed for a specific WTG.
    Returns a minimal dict with just the required data.
    """
    cat_col = find_wtg_column(columns, wtg, 'category')
    power_col = find_wtg_column(columns, wtg, 'power')
    wind_col = find_wtg_column(columns, wtg, 'wind_speed')
    air_col = find_wtg_column(columns, wtg, 'air_density')
    
    result = {
        'ts_col': ts_col,
        'cat_col': cat_col,
        'power_col': power_col,
        'wind_col': wind_col,
        'air_col': air_col,
        'timestamp': scada_data.get(ts_col, []),
        'category': scada_data.get(cat_col, []) if cat_col else [],
        'power': scada_data.get(power_col, []) if power_col else [],
        'wind_speed': scada_data.get(wind_col, []) if wind_col else [],
        'air_density': scada_data.get(air_col, []) if air_col else []
    }
    return result


def process_wtg(
    wtg_data: dict,
    wtg: str,
    alarm_df: pd.DataFrame,
    rated_power_kw: float,
    bin_minutes: float,
    test_hours: int,
    extension_hours: int,
    min_availability_pct: float,
    pr_threshold: float,
    power_curve_arrays,
    require_nominal: bool = True,
    min_nominal_hours: float = 24.0
) -> dict:
    """
    Process a single WTG to find valid run test windows.
    Memory-efficient: receives only pre-extracted columns for this WTG.
    """
    cat_col = wtg_data['cat_col']
    power_col = wtg_data['power_col']
    
    if wtg_data['ts_col'] is None:
        return {'wtg': wtg, 'status': 'FAILED', 'reason': 'No timestamp column', 'windows': [], 'viable_count': 0, 'total_alarms': 0}
    
    if cat_col is None:
        return {'wtg': wtg, 'status': 'FAILED', 'reason': f'No category column for {wtg}', 'windows': [], 'viable_count': 0, 'total_alarms': 0}
    
    if power_col is None:
        return {'wtg': wtg, 'status': 'FAILED', 'reason': f'No power column for {wtg}', 'windows': [], 'viable_count': 0, 'total_alarms': 0}
    
    # Build minimal DataFrame with only needed columns (already extracted)
    df = pd.DataFrame({
        'timestamp': pd.to_datetime(wtg_data['timestamp'], errors='coerce'),
        'category': [_normalize_category(c) for c in wtg_data['category']],
        'power': pd.to_numeric(pd.Series(wtg_data['power']), errors='coerce').astype(np.float32)
    })
    
    if wtg_data['wind_speed']:
        df['wind_speed'] = pd.to_numeric(pd.Series(wtg_data['wind_speed']), errors='coerce').astype(np.float32)
    else:
        df['wind_speed'] = np.nan
    
    if wtg_data['air_density']:
        df['air_density'] = pd.to_numeric(pd.Series(wtg_data['air_density']), errors='coerce').astype(np.float32)
    else:
        df['air_density'] = np.nan
    
    df = df.sort_values('timestamp').reset_index(drop=True)
    N = len(df)
    
    if N == 0:
        return {'wtg': wtg, 'status': 'FAILED', 'reason': 'No data', 'windows': [], 'viable_count': 0, 'total_alarms': 0}
    
    # Debug: Log power stats for this WTG
    power_series = df['power']
    power_min = float(power_series.min()) if len(power_series) > 0 else 0
    power_max = float(power_series.max()) if len(power_series) > 0 else 0
    power_mean = float(power_series.mean()) if len(power_series) > 0 else 0
    
    # Check air density availability
    air_available = df['air_density'].notna().sum()
    air_missing_count = df['air_density'].isna().sum()
    
    log(f"[Python]   {wtg}: bins={N}, power min={power_min:.0f} max={power_max:.0f} mean={power_mean:.0f} kW, air_density={air_available}/{N}")
    
    # Compute expected power from power curve
    if power_curve_arrays is not None:
        expected_power = compute_expected_power(df['wind_speed'], df['air_density'], power_curve_arrays)
        # Nominal check: actual_power / expected_power >= pr_threshold
        # No scaling on expected power - pr_threshold directly controls the comparison
        
        # Debug: check how many bins have valid expected power
        valid_expected = np.isfinite(expected_power).sum()
        exp_min = float(np.nanmin(expected_power)) if valid_expected > 0 else 0
        exp_max = float(np.nanmax(expected_power)) if valid_expected > 0 else 0
        exp_mean = float(np.nanmean(expected_power)) if valid_expected > 0 else 0
        log(f"[Python]   {wtg}: expected_power valid={valid_expected}/{N}, min={exp_min:.0f} max={exp_max:.0f} mean={exp_mean:.0f} kW")
    else:
        expected_power = np.full(N, np.nan, dtype=np.float32)
        log(f"[Python]   {wtg}: No power curve provided")
    
    # Compute nominal power flag
    power = df['power'].to_numpy()
    air_missing = df['air_density'].isna().to_numpy()
    
    # Performance ratio based nominal (where air density available)
    # pr_threshold parameter controls the comparison: actual >= expected * pr_threshold
    # e.g., pr_threshold=0.81 means bin is nominal if producing >= 81% of expected power
    perf_ratio = np.divide(power, expected_power, out=np.full(N, np.nan, dtype=np.float32), 
                           where=np.isfinite(expected_power) & (expected_power > 0))
    pr_nominal = perf_ratio >= pr_threshold  # Use parameter directly
    
    # Fallback thresholds for when expected power unavailable
    # For missing air density: 90% rated power
    threshold_missing_air = rated_power_kw * 0.90
    # For valid air density but no expected power (wind speed out of curve): use threshold % of rated
    threshold_normal = rated_power_kw * pr_threshold
    
    fallback_missing_air = np.isfinite(power) & (power >= threshold_missing_air)
    fallback_normal = np.isfinite(power) & (power >= threshold_normal)
    
    # 50% rated power for cooling check
    threshold_50pct = rated_power_kw * 0.50
    cooling_ok = np.isfinite(power) & (power >= threshold_50pct)
    
    # Combined nominal flag:
    # - If air density missing: use 90% rated power threshold
    # - Else if perf_ratio valid: use PR-based check
    # - Else (air density exists but expected power invalid): use threshold % rated
    nominal = np.where(
        air_missing, 
        fallback_missing_air,
        np.where(np.isfinite(perf_ratio), pr_nominal, fallback_normal)
    )
    
    # Debug: count nominal bins
    total_nominal = nominal.sum()
    nominal_hours_total = total_nominal * (bin_minutes / 60.0)
    log(f"[Python]   {wtg}: total nominal bins={total_nominal} ({nominal_hours_total:.1f}h), pr_nominal={pr_nominal.sum()}, fallback={fallback_missing_air.sum()}")
    
    # Category-based masks
    cats = df['category'].to_numpy()
    active = np.array([c in ACTIVE_CATEGORIES for c in cats])
    blocked = np.array([c in BLOCKED_CATEGORIES for c in cats])
    allowed_window = np.array([c in ALLOWED_WINDOW_CATEGORIES for c in cats])
    
    # Prefix sums for O(1) range queries
    ps_active = np.cumsum(active.astype(np.int32))
    ps_blocked = np.cumsum(blocked.astype(np.int32))
    ps_nominal = np.cumsum((nominal & active).astype(np.int32))
    ps_cooling = np.cumsum((cooling_ok & active).astype(np.int32))
    ps_disallowed = np.cumsum((~allowed_window).astype(np.int32))
    # For energy, replace NaN with 0 before cumsum
    power_clean = np.where(np.isfinite(power), power, 0.0)
    ps_energy = np.cumsum((power_clean * active * (bin_minutes / 60.0)).astype(np.float64))
    
    def span_sum(ps, s, e):
        return ps[e] - (ps[s-1] if s > 0 else 0)
    
    # Target bins for 72h (+ extension)
    target_bins = int(round(test_hours * 60 / bin_minutes))
    max_bins = int(round((test_hours + extension_hours) * 60 / bin_minutes)) if extension_hours > 0 else target_bins
    
    # Get alarms for this WTG
    wtg_alarms = get_alarms_for_wtg(alarm_df, wtg) if alarm_df is not None and len(alarm_df) > 0 else pd.DataFrame()
    alarm_count = len(wtg_alarms)
    
    # Minimum energy requirement: (Rated Power × Base Window Hours) ÷ 2
    min_energy_kwh = (rated_power_kw * test_hours) / 2.0
    log(f"[Python]   {wtg}: minimum energy requirement = {min_energy_kwh:.0f} kWh ({min_energy_kwh/1000:.1f} MWh)")
    
    # Helper function to check a single window
    def check_window(s, e, is_base_window=True, allow_disqualified=False):
        disallowed_count = span_sum(ps_disallowed, s, e)
        has_disqualified = disallowed_count > 0
        
        # Return None for disqualified windows unless we're collecting near-misses
        if has_disqualified and not allow_disqualified:
            return None
        
        active_bins = span_sum(ps_active, s, e)
        blocked_bins = span_sum(ps_blocked, s, e)
        nominal_bins = span_sum(ps_nominal, s, e)
        energy_kwh = span_sum(ps_energy, s, e)
        
        avail_pct = 100.0 * ((active_bins - blocked_bins) / max(1, active_bins))
        nominal_hours = nominal_bins * (bin_minutes / 60.0)
        active_hours = active_bins * (bin_minutes / 60.0)
        wall_hours = (e - s + 1) * (bin_minutes / 60.0)
        
        start_ts = df['timestamp'].iloc[s]
        end_ts = df['timestamp'].iloc[e]
        
        # Count alarms and warnings separately for comparison
        alarm_count_val, warning_count_val = count_alarms_and_warnings(wtg_alarms, start_ts, end_ts)
        
        # Get fault codes for 3-strike rule
        fault_counts = {}
        three_strike_pass = True
        violating_codes = {}
        
        if len(wtg_alarms) > 0:
            fault_counts = count_fault_codes_in_window(wtg_alarms, start_ts, end_ts)
            three_strike_pass, violating_codes = check_three_strike_rule(fault_counts)
        
        total_alarms = sum(fault_counts.values())
        
        issues = []
        
        # HARD RULE: Disqualified categories - window cannot contain disqualified bins
        if has_disqualified:
            issues.append(f"Contains {disallowed_count} disqualified bin(s)")
        
        # HARD RULE: Minimum energy check (only for base window)
        if is_base_window and energy_kwh < min_energy_kwh:
            issues.append(f"Energy {energy_kwh:.0f}kWh < {min_energy_kwh:.0f}kWh (50% rated power minimum)")
        
        if avail_pct < min_availability_pct:
            issues.append(f"Availability {avail_pct:.1f}% < {min_availability_pct}%")
        if require_nominal and nominal_hours < min_nominal_hours:
            issues.append(f"Nominal hours {nominal_hours:.1f}h < {min_nominal_hours}h")
        if not three_strike_pass:
            codes_str = ', '.join([f"{c}({n}x)" for c, n in violating_codes.items()])
            issues.append(f"3-strike violated: {codes_str}")
        
        viable = len(issues) == 0
        
        return {
            'start_idx': int(s),
            'end_idx': int(e),
            'start_time': start_ts.isoformat() if pd.notna(start_ts) else '',
            'end_time': end_ts.isoformat() if pd.notna(end_ts) else '',
            'wall_hours': float(round(wall_hours, 2)),
            'active_hours': float(round(active_hours, 2)),
            'availability_pct': float(round(avail_pct, 2)),
            'nominal_hours': float(round(nominal_hours, 2)),
            'energy_mwh': float(round(energy_kwh / 1000.0, 2)) if np.isfinite(energy_kwh) else 0.0,
            'energy_kwh': float(round(energy_kwh, 2)) if np.isfinite(energy_kwh) else 0.0,
            'total_alarms': int(total_alarms),
            'alarm_count': int(alarm_count_val),
            'warning_count': int(warning_count_val),
            'fault_codes': {str(k): int(v) for k, v in fault_counts.items()},
            'three_strike_pass': bool(three_strike_pass),
            'violating_codes': {str(k): int(v) for k, v in violating_codes.items()},
            'viable': bool(viable),
            'issues': issues,
            'disqualified_bins': int(disallowed_count)
        }
    
    # Helper to compare two viable windows: prefer fewer alarms, then fewer warnings
    def is_better_window(new_cand, current_best):
        """Return True if new_cand is better than current_best.
        Preference order: fewer alarms first, then fewer warnings."""
        if current_best is None:
            return True
        # Alarms take priority
        if new_cand['alarm_count'] < current_best['alarm_count']:
            return True
        if new_cand['alarm_count'] > current_best['alarm_count']:
            return False
        # Same alarm count - check warnings
        if new_cand['warning_count'] < current_best['warning_count']:
            return True
        return False
    
    # Search for the best viable window (prefer fewest alarms, then fewest warnings)
    # Step through data, checking each possible window position
    best_candidate = None
    best_non_viable = None  # Track best non-viable for diagnostics
    all_viable_candidates = []  # Collect all viable windows to find the best
    all_candidates = []  # Collect ALL checked windows for review
    near_miss_candidates = []  # Collect near-miss windows (with disqualified bins) for diagnostics
    step = max(1, int(1 * 60 / bin_minutes))  # 1-hour steps (finer search)
    windows_checked = 0
    
    s = 0
    
    while s < N - target_bins + 1:  # Only start positions that can fit a full window
        # Build window: find the end index that gives us target_bins active bins
        active_count = 0
        e = s - 1
        
        while e + 1 < N and active_count < target_bins:
            e += 1
            if active[e]:
                active_count += 1
        
        # Check if we have enough active bins for base window
        windows_checked += 1
        if active_count >= target_bins:
            # Check the base window first (is_base_window=True for energy check)
            base_energy_check_end = e  # Remember base window end for energy calc
            cand = check_window(s, e, is_base_window=True, allow_disqualified=False)
            
            # If no valid candidate (has disqualified bins), collect as near-miss for diagnostics
            if cand is None and len(near_miss_candidates) < 20:
                near_miss = check_window(s, e, is_base_window=True, allow_disqualified=True)
                if near_miss:
                    near_miss_candidates.append(near_miss)
            
            if cand:
                # Collect all candidates for review (limit to avoid memory issues)
                if len(all_candidates) < 100:
                    all_candidates.append(cand)
                
                if cand['viable']:
                    # Found a viable window - check if it's better than current best
                    if is_better_window(cand, best_candidate):
                        best_candidate = cand
                        log(f"[Python]   Found viable window at idx {s}: nominal={cand['nominal_hours']:.1f}h, alarms={cand['alarm_count']}, warnings={cand['warning_count']}")
                    
                    # If this window has no alarms and no warnings, it's optimal - stop searching
                    if cand['alarm_count'] == 0 and cand['warning_count'] == 0:
                        log(f"[Python]   Optimal window found (no alarms/warnings) - stopping search")
                        all_candidates.append(cand)  # Make sure optimal is included
                        break
                    
                    # Continue searching for better windows
                    all_viable_candidates.append(cand)
                
                # If base window passes hard rules but doesn't have enough nominal, try extending
                # Only extend if base window passes: no disqualified bins and meets energy requirement
                base_passes_hard_rules = ('Energy' not in ' '.join(cand.get('issues', [])) and 
                                          cand is not None)
                
                if base_passes_hard_rules and require_nominal and cand['nominal_hours'] < min_nominal_hours and max_bins > target_bins:
                    # Try extending the window to get more nominal hours
                    # Use minimum extension needed to reach target nominal hours
                    extended_e = e
                    extended_active = active_count
                    current_nominal = cand['nominal_hours']
                    
                    # Extend incrementally, checking if we've reached the nominal target
                    while extended_e + 1 < N and extended_active < max_bins:
                        extended_e += 1
                        if active[extended_e]:
                            extended_active += 1
                        
                        # Check if this bin adds nominal time
                        if nominal[extended_e] and active[extended_e]:
                            current_nominal += (bin_minutes / 60.0)
                            
                            # If we've reached the target, stop extending
                            if current_nominal >= min_nominal_hours:
                                break
                    
                    if extended_active > active_count:
                        # Extended window - is_base_window=False since energy was checked on base
                        ext_cand = check_window(s, extended_e, is_base_window=False)
                        if ext_cand:
                            # Mark as extended window
                            ext_cand['is_extended'] = True
                            ext_cand['extension_hours'] = round((extended_active - active_count) * (bin_minutes / 60.0), 2)
                            
                            # Collect extended candidates too
                            if len(all_candidates) < 100:
                                all_candidates.append(ext_cand)
                            
                            if ext_cand['viable']:
                                if is_better_window(ext_cand, best_candidate):
                                    best_candidate = ext_cand
                                    log(f"[Python]   Found viable EXTENDED window at idx {s}: active={ext_cand['active_hours']:.1f}h, nominal={ext_cand['nominal_hours']:.1f}h, alarms={ext_cand['alarm_count']}, warnings={ext_cand['warning_count']}")
                                
                                # If this extended window has no alarms and no warnings, it's optimal
                                if ext_cand['alarm_count'] == 0 and ext_cand['warning_count'] == 0:
                                    log(f"[Python]   Optimal extended window found (no alarms/warnings) - stopping search")
                                    break
                                
                                all_viable_candidates.append(ext_cand)
                            else:
                                # Use extended as non-viable candidate if it has more nominal hours
                                if best_non_viable is None or ext_cand['nominal_hours'] > best_non_viable['nominal_hours']:
                                    best_non_viable = ext_cand
                
                # Track best non-viable (by nominal hours) if this window isn't viable
                if not cand['viable']:
                    if best_non_viable is None or cand['nominal_hours'] > best_non_viable['nominal_hours']:
                        best_non_viable = cand
        
        # Move to next start position
        s += step
    
    # Sort all_candidates by: viable first, then by (fewer alarms, fewer warnings, more nominal hours)
    def window_sort_key(w):
        return (
            0 if w['viable'] else 1,  # Viable first
            w.get('alarm_count', 999),  # Fewer alarms
            w.get('warning_count', 999),  # Fewer warnings
            -w.get('nominal_hours', 0),  # More nominal hours
            -w.get('energy_mwh', 0)  # More energy
        )
    all_candidates.sort(key=window_sort_key)
    
    # Also sort near-miss candidates by fewest disqualified bins first
    near_miss_candidates.sort(key=lambda w: (w.get('disqualified_bins', 999), len(w.get('issues', [])), -w.get('nominal_hours', 0)))
    
    viable_found = len(all_viable_candidates) + (1 if best_candidate and best_candidate['viable'] else 0)
    log(f"[Python]   {wtg}: checked {windows_checked} windows, found {viable_found} viable, collected {len(all_candidates)} candidates, {len(near_miss_candidates)} near-misses")
    
    if best_candidate and best_candidate['viable']:
        log(f"[Python]   {wtg}: BEST viable window - alarms={best_candidate['alarm_count']}, warnings={best_candidate['warning_count']}, energy={best_candidate['energy_mwh']:.1f}MWh")
    elif near_miss_candidates:
        nm = near_miss_candidates[0]
        log(f"[Python]   {wtg}: BEST near-miss - {nm.get('disqualified_bins', 0)} disqualified bins, "
            f"nominal={nm.get('nominal_hours', 0):.1f}h, avail={nm.get('availability_pct', 0):.1f}%, "
            f"energy={nm.get('energy_mwh', 0):.1f}MWh")
    
    # Debug: analyze the best window to understand nominal calculation
    if best_non_viable and not (best_candidate and best_candidate['viable']):
        bw = best_non_viable
        s, e = bw['start_idx'], bw['end_idx']
        window_nominal = nominal[s:e+1]
        window_active = active[s:e+1]
        window_pr = perf_ratio[s:e+1]
        window_power = power[s:e+1]
        window_expected = expected_power[s:e+1]
        
        # Count how PR values are distributed in this window
        pr_valid = np.isfinite(window_pr)
        pr_ge_1 = (window_pr >= 1.0) & pr_valid
        pr_ge_099 = (window_pr >= 0.99) & pr_valid
        pr_ge_097 = (window_pr >= 0.97) & pr_valid
        
        log(f"[Python]   Best window PR distribution: valid={pr_valid.sum()}, >=1.0={pr_ge_1.sum()}, >=0.99={pr_ge_099.sum()}, >=0.97={pr_ge_097.sum()}")
        log(f"[Python]   Best window: power mean={np.nanmean(window_power):.0f}, expected mean={np.nanmean(window_expected):.0f}")
    
    # Use best viable, or fall back to best non-viable for diagnostics
    if best_candidate is None:
        best_candidate = best_non_viable
    
    # Build results
    candidates = [best_candidate] if best_candidate else []
    viable_count = 1 if (best_candidate and best_candidate['viable']) else 0
    
    # Generate detailed failure reason if no viable windows
    failure_details = []
    if viable_count == 0 and candidates:
        # Analyze best non-viable candidate to explain why
        best_attempt = candidates[0]
        # Check for energy failure
        if best_attempt.get('energy_kwh', 0) < min_energy_kwh:
            failure_details.append(f"Energy {best_attempt.get('energy_kwh', 0):.0f}kWh < {min_energy_kwh:.0f}kWh (50% rated power minimum)")
        if best_attempt['availability_pct'] < min_availability_pct:
            failure_details.append(f"Availability {best_attempt['availability_pct']:.1f}% < {min_availability_pct}% required")
        if require_nominal and best_attempt['nominal_hours'] < min_nominal_hours:
            failure_details.append(f"Nominal hours {best_attempt['nominal_hours']:.1f}h < {min_nominal_hours}h required")
        if not best_attempt['three_strike_pass']:
            codes = [f"{c}({n}x)" for c, n in best_attempt['violating_codes'].items()]
            failure_details.append(f"3-strike rule violated: {', '.join(codes)}")
        if best_attempt.get('alarm_count', 0) > 0:
            failure_details.append(f"{best_attempt.get('alarm_count', 0)} alarms, {best_attempt.get('warning_count', 0)} warnings in window")
    elif viable_count == 0:
        # No candidates at all - check if we have near-miss candidates to report on
        if near_miss_candidates:
            # Sort near-miss by fewest issues first
            near_miss_candidates.sort(key=lambda w: (w.get('disqualified_bins', 999), len(w.get('issues', [])), -w.get('nominal_hours', 0)))
            best_near_miss = near_miss_candidates[0]
            failure_details.append(f"All {test_hours}h windows contain disqualified bins (e.g., Manufacturer/Unscheduled maintenance)")
            failure_details.append(f"Best near-miss: {best_near_miss.get('disqualified_bins', 0)} disqualified bins, "
                                   f"{best_near_miss.get('nominal_hours', 0):.1f}h nominal, "
                                   f"{best_near_miss.get('availability_pct', 0):.1f}% availability")
            # Add near-miss candidates to all_candidates for display
            all_candidates = near_miss_candidates[:20]
        else:
            failure_details.append(f"No valid {test_hours}h windows found in data")
        failure_details.append(f"Total alarms for WTG: {alarm_count}")
    
    failure_reason = '; '.join(failure_details) if failure_details else 'No windows meet all requirements'
    
    # If no regular candidates but we have near-misses, use those for diagnostics
    if not candidates and near_miss_candidates:
        candidates = [near_miss_candidates[0]]  # Include best near-miss
    
    # Clean up
    del df
    gc.collect()
    
    return {
        'wtg': wtg,
        'status': 'OK' if viable_count > 0 else 'NO_VIABLE_WINDOW',
        'reason': '' if viable_count > 0 else failure_reason,
        'windows': candidates,
        'all_candidates': all_candidates[:20],  # Top 20 candidates for review
        'near_miss_candidates': near_miss_candidates[:10],  # Top 10 near-miss windows
        'viable_count': viable_count,
        'total_alarms': alarm_count
    }


# =============================================================================
# MAIN ANALYSIS ENTRY POINT
# =============================================================================

def run_analysis(
    scada_data: dict,
    scada_columns: list,
    alarm_data: dict = None,
    alarm_columns: list = None,
    power_curve_text: str = None,
    selected_wtgs: list = None,
    rated_power_kw: float = 4200.0,
    bin_minutes: float = 10.0,
    test_hours: int = 72,
    extension_hours: int = 24,
    min_availability_pct: float = 96.0,
    pr_threshold: float = 0.90,
    require_nominal: bool = True,
    min_nominal_hours: float = 24.0
) -> dict:
    """
    Main entry point for run test analysis.
    Processes WTGs sequentially to minimize memory usage.
    """
    global cached_analysis_data
    
    log(f"[Python] Starting Enhanced Run Test Analysis")
    log(f"[Python] Parameters: rated_power={rated_power_kw}kW, bin={bin_minutes}min, "
          f"test={test_hours}h, extension={extension_hours}h, avail>={min_availability_pct}%, "
          f"pr_threshold={pr_threshold:.0%}, nominal_req={min_nominal_hours}h")
    
    # Parse power curve if provided
    power_curve_arrays = None
    if power_curve_text and power_curve_text.strip():
        try:
            curve_df = parse_power_curve_text(power_curve_text)
            power_curve_arrays = prepare_power_curve(curve_df)
            log(f"[Python] Power curve loaded: {len(curve_df)} wind speed bins, "
                  f"{len(curve_df.columns)-1} air density columns")
        except Exception as e:
            log(f"[Python] Warning: Could not parse power curve: {e}")
    
    # Parse alarm log if provided
    alarm_df = None
    if alarm_data and alarm_columns:
        try:
            alarm_df = parse_alarm_log(alarm_data, alarm_columns)
            log(f"[Python] Alarm log loaded: {len(alarm_df)} events")
        except Exception as e:
            log(f"[Python] Warning: Could not parse alarm log: {e}")
            alarm_df = pd.DataFrame()
    else:
        alarm_df = pd.DataFrame()
    
    # Detect WTGs
    all_wtgs = detect_wtgs(scada_columns)
    log(f"[Python] Detected {len(all_wtgs)} WTGs: {', '.join(all_wtgs)}")
    
    # Filter to selected WTGs if specified
    if selected_wtgs and len(selected_wtgs) > 0:
        wtgs_to_process = [w for w in selected_wtgs if w in all_wtgs]
    else:
        wtgs_to_process = all_wtgs
    
    log(f"[Python] Processing {len(wtgs_to_process)} WTGs...")
    
    # Find timestamp column once
    ts_col = find_timestamp_column(scada_columns)
    
    # Process each WTG sequentially
    results = []
    
    for i, wtg in enumerate(wtgs_to_process):
        # Extract only the columns needed for this WTG
        wtg_data = extract_wtg_columns(scada_data, scada_columns, wtg, ts_col)
        
        result = process_wtg(
            wtg_data=wtg_data,
            wtg=wtg,
            alarm_df=alarm_df,
            rated_power_kw=rated_power_kw,
            bin_minutes=bin_minutes,
            test_hours=test_hours,
            extension_hours=extension_hours,
            min_availability_pct=min_availability_pct,
            pr_threshold=pr_threshold,
            power_curve_arrays=power_curve_arrays,
            require_nominal=require_nominal,
            min_nominal_hours=min_nominal_hours
        )
        results.append(result)
        
        # Delete extracted data and force garbage collection
        del wtg_data
        gc.collect()
    
    # Free memory - delete alarm dataframe and power curve now that we're done
    if alarm_df is not None:
        del alarm_df
    if power_curve_arrays is not None:
        del power_curve_arrays
    gc.collect()
    
    # Summary
    viable_wtgs = sum(1 for r in results if r['viable_count'] > 0)
    total_viable = sum(r['viable_count'] for r in results)
    log(f"[Python] Analysis complete: {viable_wtgs}/{len(wtgs_to_process)} WTGs have viable windows, "
          f"{total_viable} total viable windows found")
    
    # Cache for Excel export
    cached_analysis_data = {
        'results': results,
        'parameters': {
            'rated_power_kw': rated_power_kw,
            'bin_minutes': bin_minutes,
            'test_hours': test_hours,
            'extension_hours': extension_hours,
            'min_availability_pct': min_availability_pct,
            'pr_threshold': pr_threshold
        }
    }
    
    return {
        'results': results,
        'summary': {
            'wtgs_processed': len(wtgs_to_process),
            'wtgs_with_viable': viable_wtgs,
            'total_viable_windows': total_viable
        }
    }


def get_window_bins(
    scada_data: dict,
    scada_columns: list,
    wtg: str,
    start_idx: int,
    end_idx: int,
    power_curve_text: str = None,
    rated_power_kw: float = 4500.0,
    bin_minutes: float = 10.0,
    pr_threshold: float = 0.81,
    offset: int = 0,
    limit: int = 10
) -> dict:
    """
    Get detailed bin-level data for a specific WTG window.
    Returns bins sorted by power (descending) with pagination support.
    
    Args:
        scada_data: SCADA data dictionary
        scada_columns: Column names
        wtg: WTG identifier (e.g., 'A01')
        start_idx: Window start index
        end_idx: Window end index
        power_curve_text: Power curve text for expected power calculation
        rated_power_kw: Rated power in kW
        bin_minutes: Minutes per bin
        pr_threshold: Performance ratio threshold
        offset: Pagination offset
        limit: Number of bins to return
    
    Returns:
        Dictionary with bins array and metadata
    """
    ts_col = find_timestamp_column(scada_columns)
    
    # Extract WTG columns
    wtg_data = extract_wtg_columns(scada_data, scada_columns, wtg, ts_col)
    
    if wtg_data['ts_col'] is None or wtg_data['power_col'] is None:
        return {'error': f'Missing columns for {wtg}', 'bins': [], 'total': 0}
    
    # Build DataFrame for the window
    df = pd.DataFrame({
        'timestamp': pd.to_datetime(wtg_data['timestamp'], errors='coerce'),
        'category': [_normalize_category(c) for c in wtg_data['category']],
        'power': pd.to_numeric(pd.Series(wtg_data['power']), errors='coerce').astype(np.float32)
    })
    
    if wtg_data['wind_speed']:
        df['wind_speed'] = pd.to_numeric(pd.Series(wtg_data['wind_speed']), errors='coerce').astype(np.float32)
    else:
        df['wind_speed'] = np.nan
    
    if wtg_data['air_density']:
        df['air_density'] = pd.to_numeric(pd.Series(wtg_data['air_density']), errors='coerce').astype(np.float32)
    else:
        df['air_density'] = np.nan
    
    df = df.sort_values('timestamp').reset_index(drop=True)
    
    # Parse power curve if provided
    power_curve_arrays = None
    if power_curve_text and power_curve_text.strip():
        try:
            curve_df = parse_power_curve_text(power_curve_text)
            power_curve_arrays = prepare_power_curve(curve_df)
        except:
            pass
    
    # Compute expected power
    if power_curve_arrays is not None:
        expected_power = compute_expected_power(df['wind_speed'], df['air_density'], power_curve_arrays)
    else:
        expected_power = np.full(len(df), np.nan, dtype=np.float32)
    
    # Slice to window
    window_df = df.iloc[start_idx:end_idx+1].copy()
    window_expected = expected_power[start_idx:end_idx+1]
    
    # Compute performance ratio
    power_arr = window_df['power'].to_numpy()
    perf_ratio = np.divide(power_arr, window_expected, 
                           out=np.full(len(power_arr), np.nan, dtype=np.float32),
                           where=np.isfinite(window_expected) & (window_expected > 0))
    
    # Determine nominal status
    nominal = perf_ratio >= pr_threshold
    
    # Build bin records
    bins = []
    for i, (idx, row) in enumerate(window_df.iterrows()):
        bins.append({
            'idx': int(idx),
            'timestamp': row['timestamp'].isoformat() if pd.notna(row['timestamp']) else '',
            'category': row['category'],
            'power_kw': round(float(row['power']), 1) if np.isfinite(row['power']) else None,
            'wind_speed': round(float(row['wind_speed']), 2) if np.isfinite(row['wind_speed']) else None,
            'air_density': round(float(row['air_density']), 4) if np.isfinite(row['air_density']) else None,
            'expected_power': round(float(window_expected[i]), 1) if np.isfinite(window_expected[i]) else None,
            'perf_ratio': round(float(perf_ratio[i]), 3) if np.isfinite(perf_ratio[i]) else None,
            'is_nominal': bool(nominal[i]) if np.isfinite(perf_ratio[i]) else False,
            'is_active': row['category'] in ACTIVE_CATEGORIES
        })
    
    # Sort by power descending (top power bins first)
    bins.sort(key=lambda b: b['power_kw'] if b['power_kw'] is not None else -999999, reverse=True)
    
    total = len(bins)
    
    # Apply pagination
    paginated_bins = bins[offset:offset+limit]
    
    return {
        'wtg': wtg,
        'start_idx': start_idx,
        'end_idx': end_idx,
        'total_bins': total,
        'offset': offset,
        'limit': limit,
        'has_more': offset + limit < total,
        'bins': paginated_bins
    }


def generate_excel() -> bytes:
    """Generate Excel export from cached analysis data."""
    global output_excel_bytes, cached_analysis_data
    
    if cached_analysis_data is None:
        raise ValueError("No analysis data available. Run analysis first.")
    
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # In Pyodide, openpyxl needs to be installed via micropip
        raise ImportError("openpyxl required for Excel export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    # Summary headers
    headers = ['WTG', 'Status', 'Viable Windows', 'Total Alarms', 'Best Start', 'Best End', 
               'Availability (%)', 'Nominal Hours', 'Energy (MWh)', 'Issues']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
    
    # Data rows
    results = cached_analysis_data['results']
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result['wtg'])
        ws.cell(row=row_idx, column=2, value=result['status'])
        ws.cell(row=row_idx, column=3, value=result['viable_count'])
        ws.cell(row=row_idx, column=4, value=result.get('total_alarms', 0))
        
        if result['windows']:
            best = result['windows'][0]
            ws.cell(row=row_idx, column=5, value=best['start_time'])
            ws.cell(row=row_idx, column=6, value=best['end_time'])
            ws.cell(row=row_idx, column=7, value=best['availability_pct'])
            ws.cell(row=row_idx, column=8, value=best['nominal_hours'])
            ws.cell(row=row_idx, column=9, value=best['energy_mwh'])
            ws.cell(row=row_idx, column=10, value='; '.join(best.get('issues', [])))
    
    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    output_excel_bytes = buffer.getvalue()
    
    return output_excel_bytes


def get_excel_bytes() -> bytes:
    """Return cached Excel bytes."""
    global output_excel_bytes
    return output_excel_bytes


# =============================================================================
# INCREMENTAL PROCESSING FOR REAL-TIME UI UPDATES
# =============================================================================

# Global state for incremental processing
_analysis_state = {}

def init_analysis(scada_data, scada_columns, alarm_data, alarm_columns, power_curve_text, 
                  selected_wtgs, params):
    """Initialize analysis state - call once before processing WTGs."""
    global _analysis_state
    
    log(f"[Python] Starting Enhanced Run Test Analysis")
    log(f"[Python] Parameters: rated_power={params.get('rated_power_kw', 4200)}kW")
    
    # Parse power curve
    power_curve_arrays = None
    if power_curve_text and power_curve_text.strip():
        try:
            curve_df = parse_power_curve_text(power_curve_text)
            power_curve_arrays = prepare_power_curve(curve_df)
            log(f"[Python] Power curve loaded: {len(curve_df)} wind speed bins")
        except Exception as e:
            log(f"[Python] Warning: Could not parse power curve: {e}")
    
    # Parse alarm log
    alarm_df = None
    if alarm_data and alarm_columns:
        try:
            alarm_df = parse_alarm_log(alarm_data, alarm_columns)
            log(f"[Python] Alarm log loaded: {len(alarm_df)} events")
        except Exception as e:
            log(f"[Python] Warning: Could not parse alarm log: {e}")
            alarm_df = pd.DataFrame()
    else:
        alarm_df = pd.DataFrame()
    
    # Find timestamp column
    ts_col = find_timestamp_column(scada_columns)
    
    # Determine WTGs to process
    all_wtgs = detect_wtgs(scada_columns)
    if selected_wtgs and len(selected_wtgs) > 0:
        wtgs_to_process = [w for w in selected_wtgs if w in all_wtgs]
    else:
        wtgs_to_process = all_wtgs
    
    log(f"[Python] Processing {len(wtgs_to_process)} WTGs...")
    
    _analysis_state = {
        'scada_data': scada_data,
        'scada_columns': scada_columns,
        'alarm_df': alarm_df,
        'power_curve_arrays': power_curve_arrays,
        'ts_col': ts_col,
        'wtgs': wtgs_to_process,
        'params': params,
        'results': [],
        'current_idx': 0
    }
    
    return {'wtgs': wtgs_to_process, 'count': len(wtgs_to_process)}


def process_next_wtg():
    """Process the next WTG in queue. Returns result dict with 'done' flag."""
    global _analysis_state, cached_analysis_data
    
    state = _analysis_state
    idx = state['current_idx']
    wtgs = state['wtgs']
    
    if idx >= len(wtgs):
        return {'done': True, 'result': None}
    
    wtg = wtgs[idx]
    params = state['params']
    
    # Extract columns for this WTG
    wtg_data = extract_wtg_columns(state['scada_data'], state['scada_columns'], 
                                    wtg, state['ts_col'])
    
    # Process
    result = process_wtg(
        wtg_data=wtg_data,
        wtg=wtg,
        alarm_df=state['alarm_df'],
        rated_power_kw=params.get('rated_power_kw', 4200),
        bin_minutes=params.get('bin_minutes', 10),
        test_hours=params.get('test_hours', 72),
        extension_hours=params.get('extension_hours', 24),
        min_availability_pct=params.get('min_availability_pct', 96),
        pr_threshold=params.get('pr_threshold', 0.81),  # Default 81% matches old tool
        power_curve_arrays=state['power_curve_arrays'],
        require_nominal=params.get('require_nominal', True),
        min_nominal_hours=params.get('min_nominal_hours', 24)
    )
    
    state['results'].append(result)
    state['current_idx'] = idx + 1
    
    # Log result
    status = "PASS" if result['viable_count'] > 0 else "FAIL"
    if result['windows']:
        best = result['windows'][0]
        strike_info = "3x:OK" if best['three_strike_pass'] else f"3x:FAIL({len(best['violating_codes'])})"
        log(f"[Python]   {wtg}: {status} | "
              f"active={best['active_hours']:.1f}h, "
              f"nominal={best['nominal_hours']:.1f}h, "
              f"avail={best['availability_pct']:.1f}%, "
              f"energy={best['energy_mwh']:.1f}MWh, "
              f"{strike_info}, "
              f"alarms={best['total_alarms']}")
    else:
        log(f"[Python]   {wtg}: {status} | {result.get('reason', 'No windows')}")
    
    # Clean up
    del wtg_data
    gc.collect()
    
    return {'done': False, 'result': result, 'wtg': wtg, 'index': idx}


def finalize_analysis():
    """Finalize analysis and return full results."""
    global _analysis_state, cached_analysis_data
    
    state = _analysis_state
    results = state['results']
    
    # Clean up
    if state.get('alarm_df') is not None:
        del state['alarm_df']
    if state.get('power_curve_arrays') is not None:
        del state['power_curve_arrays']
    gc.collect()
    
    viable_wtgs = sum(1 for r in results if r['viable_count'] > 0)
    total_viable = sum(r['viable_count'] for r in results)
    
    log(f"[Python] Analysis complete: {viable_wtgs}/{len(state['wtgs'])} WTGs have viable windows")
    
    # Cache for Excel
    cached_analysis_data = {
        'results': results,
        'parameters': state['params']
    }
    
    _analysis_state = {}
    
    return {
        'results': results,
        'summary': {
            'wtgs_processed': len(state['wtgs']),
            'wtgs_with_viable': viable_wtgs,
            'total_viable_windows': total_viable
        }
    }
